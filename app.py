from flask import Flask, render_template, request, jsonify
import folium
from folium.plugins import Draw
from geopy.geocoders import Nominatim
import webbrowser
from shapely.geometry import Polygon
from shapely.geometry import Polygon, Point
from pyproj import Proj, transform
import random
import math
import pyproj
import geojson
import os
import matplotlib.pyplot as plt
from flask import send_file
import warnings
import tempfile
import html
import json
import time
import numpy as np
from random import shuffle
import pandas as pd
import openpyxl
from datetime import datetime
import re
from zipfile import ZipFile
import csv
import zipfile

# Ignore the specific warnings
warnings.filterwarnings("ignore", category=FutureWarning, message="This function is deprecated")

#username = input('Key in Straits Construction username: ')
username = 'radom'

clockfile = rf'c:\Users\{username}\OneDrive - Straits Construction Singapore Pte Ltd\120 Internship RP\10 Assignment\101_Resources\Data\MClock Record_202406.xlsx'


app = Flask(__name__)

def create_temp_dirs(): 
    global UPLOAD_FOLDER 
    global OUTPUT_FOLDER 
    global ZIP_FOLDER 
    global employeedict
    global coordinatesdf
    global circlegenerateddf
    global replacecontent
    global searchcontent
    global content

    #initialize the variables

    content = ''
    replacecontent = ''
    searchcontent = ''
    employeedict = {}
    coordinatesdf = pd.DataFrame()
    circlegenerateddf = pd.DataFrame()

    OUTPUT_FOLDER = tempfile.TemporaryDirectory() #Creates a temporary directory to store the uploaded excel files
    ZIP_FOLDER = tempfile.TemporaryDirectory() #Creates a temporary directory to store the output which would be the txt and the csv
    UPLOAD_FOLDER = tempfile.TemporaryDirectory() #Creates a temporary directory to store the zip file of the output

    
create_temp_dirs()

def deletefiles(): #to clear up the directories
    UPLOAD_FOLDER.cleanup()
    OUTPUT_FOLDER.cleanup()
    ZIP_FOLDER.cleanup()
    create_temp_dirs()


@app.route('/')
def base():
    '''map = folium.Map(
        location=[1.3521, 103.8198],
        zoom_start=12,
        max_bounds=True,
        control_scale=True
    )
    map = Draw.add_to(map)

    map.save('static/map.html')''' # To create the folium map.html but dont need it anymore because there are some configurations that cannot be made in here

    global projectplace
    global locationlist
    global projectplacecoords
    global projectplacecontent
    global chooseemployee

    projectplacecontent = ''

    excel_file = 'Project Sites.xlsx' #<= use this  when you are running in vscode

    #excel_file = os.path.join(sys._MEIPASS, 'Project Sites.xlsx') #this is when you are converting into a .exe file using pyinstaller
    projectplacecoords = []
    locationlist = []
    projectplace = []

    inputfiledf = pd.read_excel(excel_file)
    projectplacecontent += '=' * 50
    projectplacecontent += '\n\nPROJECT SITES\n\n'
    projectplacecontent += '=' * 50

    #to get the projectsite data that is stored in the same directory as this script
    if 'Location' in inputfiledf.columns:
        columnname = 'Location'
        postaladdress = 'Address'
    else:
        columnname = 'Project Code'
        postaladdress = 'Postal Code'

        for latitude, longitude, location, address in zip(inputfiledf['Latitude'], inputfiledf['Longitude'], inputfiledf[columnname], inputfiledf[postaladdress]):
            #gets the coordinates, project site name and the address
            if (longitude is not None and latitude is not None) and (longitude != 0 and latitude != 0):
                center = [longitude, latitude]
                projectplacecoords.append(center)
                locationlist.append(location)
                if address == None:
                    address = 'Unspecified'
                else:
                    address = str(address)
                projectplace.append({'Location':location,'Address': (address), 'Center':[longitude, latitude], 'shapetype': 'circle'})

                projectplacecontent += f'\n{location}\n{postaladdress}: {address}\nLongitude: {longitude} Latitude: {latitude}\n'
    

    chooseemployee = []

    #to get the employeename so user can filter to just getting the data for that employee
    if os.path.exists(clockfile):
        wb = openpyxl.load_workbook(clockfile)
        sheetnames = wb.sheetnames
        for sheet in sheetnames:
            inputfiledf = pd.read_excel(clockfile, sheet_name=sheet)
            for latitude, longitude, employeenamecode in zip(inputfiledf['Latitude'], inputfiledf['Longitude'], inputfiledf['EmployeeCodeName']):
                #checks if the coordinates are empty and get the employeenamecode
                if (longitude != None and latitude != None) and (longitude != 0 and latitude != 0):
                    chooseemployee.append(employeenamecode)

    chooseemployee = list(set(chooseemployee))

    print(locationlist)
    #passes back the choosemployee to insert in the dropdown box to let the user to select the employee before processing and also initializing the index.html
    return render_template('index.html', chooseemployee=chooseemployee)

@app.route('/reset', methods=['POST']) #to remove the uploaded file
def reset():
    deletefiles()
    return

@app.route('/search_location', methods=['POST']) #when user is entered a location in the search bar
def search_location():
    location_name = request.form['location_name']
    geolocator = Nominatim(user_agent="straitsconstruction") #this is to let the service provider to check the traffic 
    location = geolocator.geocode(location_name) #get the lattitude and the longitude of the location name

    if location:
        if 1.16 <= location.latitude <= 1.5607 and 103.502 <= location.longitude <= 104.14: #check if it is in the singapore boundary
            return jsonify({
                'success': True,
                'latitude': location.latitude,
                'longitude': location.longitude,
                'location_name': location_name,
                'zoom_start' : 17
            })
        else:
            return jsonify({'success': False, 'error': "Location must be within Singapore and it must be the location's full name."}), 400
    else:
        return jsonify({'success': False, 'error': "Location cannot be found."}), 400


@app.route('/calculate_area', methods=['POST']) #this is the circle generator
def calculate_area():

    deletefiles()
    global searchcontent
    global printarea
    global number
    global circlegenerateddf
    global replacecontent
    global content

    replacecontent = ''

    circlegenerateddf = pd.DataFrame()
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'No coordinates provided'}), 400
    searchcontent = ''
    coordinates = data['coordinate']   # get the coordinates of the polygon
    coordinates = coordinates.replace('[', '')
    coordinatessplit = coordinates.split('],')
    coordinateslist = []
    for i in coordinatessplit: #it is in string cause of the json.stringify in index.html
        spliti = i.split(',')
        floater = [float(spliti[0]), float(spliti[1])]
        coordinateslist.append(floater)
    
    polygon = Polygon(coordinateslist) # format [[x1, y1], [x2, y2], ...]
    #area = polygon.area #originally degree square

    #define projection systems
    geodetic = Proj(init='epsg:4326')  # WGS84
    utm = Proj(init='epsg:32648')  # UTM Zone 48N (Singapore)

    #project the polygon into UTM projection
    projected_coords = [transform(geodetic, utm, lon, lat) for lon, lat in polygon.exterior.coords]
    projected_polygon = Polygon(projected_coords)

    #calculate the area of the projected polygon (in square meters)
    area = projected_polygon.area
    printarea = '{:.2f}'.format(area)

    radius = float(data['radius'])
    circlenumber = int(data['circlenumber'])    

    circle_positions = []
    existing_circles=[]

    def circles_overlap(circle1, circle2): #check if it overlapped more than its radius
        if circle2 != []:
            checkall = False
            for circle in circle2:
                distance = Point(circle1['x'], circle1['y']).distance(Point(circle['x'], circle['y'])) #check the distance of the two circle center point
                if distance < circle1['radius']: #if the distance is smaller than the radius it will not be accepted because it overlap alot
                    return False
                else:
                    checkall = True
            
            if checkall == True:
                return True
            
        else: #so if the list is empty it will put the first circle then compare the rest if needed
            return True

    def circle_inside_polygon(coordinates): #check whether the center of the circle is inside the polygon
        number = 0
        for coordinate in coordinates:
            x = coordinate[0]
            y = coordinate[1]
            
            if polygon.contains(Point(x, y)):
                #number += 1
                return True
        
        

    def circle_coordinates(x_center, y_center, radius, num_points=100): #find the coordinates around the circle's circumference and put it as 100 points
        coordinates = []
        for i in range(num_points):
            theta = 2 * math.pi * i / num_points
            x = x_center + radius * math.cos(theta)
            y = y_center + radius * math.sin(theta)
            coordinates.append((x, y))
        return coordinates

    def circle_to_geojson(x_center, y_center, radius, num_points=100):
        coordinates = circle_coordinates(x_center, y_center, radius, num_points)
        polygon = geojson.Polygon([coordinates])
        return polygon


    max_time_limit = 30 #give a 30 seconds run if it exceeds it will produce what it could get
    

    # Start the timer
    start_time = time.time()

    def finduncoveredpoints(polygon, circle_positions,radius, grid_density=50): #creates a meshgrid for the polygon with 50 points to check if the circles have covered the points of the polygon
        min_x, min_y, max_x, max_y = polygon.bounds
        grid_x, grid_y = np.meshgrid(np.linspace(min_x, max_x, grid_density), np.linspace(min_y, max_y, grid_density))
        grid_points = np.vstack([grid_x.ravel(), grid_y.ravel()]).T
        
        # Check which points are inside the polygon
        inside_polygon = np.array([polygon.contains(Point(x, y)) for x, y in grid_points])
        covered_by_circles = np.zeros(grid_points.shape[0], dtype=bool)
        for circle in circle_positions:
            circle_center = circle['center']
            distances = np.sqrt((grid_points[:, 0] - circle_center[0]) ** 2 + (grid_points[:, 1] - circle_center[1]) ** 2)
            covered_by_circles = covered_by_circles | (distances <= radius)
        
        # Points that are inside the polygon but not covered by circles
        uncovered_points = grid_points[inside_polygon & ~covered_by_circles]
        formatted_points = uncovered_points.tolist()

        return formatted_points

    outofbounds = False
    maxcircle = False

    if data['isChecked'] == True:
        circlenumber = 1

    for _ in range(circlenumber):
        #check if exceed 30 seconds
        if time.time() - start_time > max_time_limit:
            if circle_positions == []:
                outofbounds = True
            else:
                maxcircle = True
            break
        while True:
            if time.time() - start_time > max_time_limit:
                if circle_positions == []:
                    outofbounds = True
                else:
                    maxcircle = True
                break

            min_x, max_y, max_x, min_y= polygon.bounds
            #randomly generate a coordinate within the polygon boundaries
            x = random.uniform(min_x, max_x)
            y = random.uniform(min_y, max_y)

            center = [x, y]

            if polygon.contains(Point(x, y)):
                source_crs = pyproj.Proj(init='epsg:3857')

                #to turn it into a degree metric from web mercator
                center_x, center_y = pyproj.transform(source_crs, geodetic, y, x)
                radius_deg = pyproj.transform(source_crs, geodetic, radius+1.34, 0)[0] - center_x

                new_circle = {'x': x, 'y': y, 'radius': radius_deg}

                #getting the coordinates of the circumference and make it as a "polygon variable"
                circle_geojson = circle_to_geojson(x, y, radius_deg)

                #making it a complete polygon by mentioning the first coordinates at the end of the list
                coordinates = circle_geojson['coordinates'][0]
                coordinatesfirst = coordinates[0]
                coordinates.append(coordinatesfirst)


                if circles_overlap(new_circle, existing_circles):
                    circle_positions.append({"center":center,"coordinates":coordinates}) #store coordinates
                    existing_circles.append(new_circle)
                    break

            if time.time() - start_time > max_time_limit:
                if circle_positions == []:
                    outofbounds = True
                else:
                    maxcircle = True
                break


    circlecoordinateslist = []
    number = 0
    i = 0
    notify = ''
    #to create a tet document for content of the coordinates of the circle, radius and area size
    content = f'Area size (Land): {printarea}m²\n'
    content += f'Circle Radius: {radius}m\n'
    if outofbounds == True: #this is for the circle that is out of bounds meaning that their diameter is bigger than the land
        number = 1
        sumx = 0
        sumy = 0
        n = len(coordinateslist)

        while i < n:
            sumx += coordinateslist[i][0]
            sumy += coordinateslist[i][1]
            i += 1
        
        centerX = sumx / n
        centerY = sumy / n

        center = [centerX, centerY]
        longitude = '{:.6f}'.format(centerX)
        latitude = '{:.6f}'.format(centerY)

        circlecoordinateslist.append(center)
        

        content += f'\nCircle {number}\nCenter:\nLongitude: {longitude} Latitude: {latitude}\n'
        notify = 'Circle is out of bounds'


        circlegenerateddf = pd.DataFrame(circlecoordinateslist, columns=['Longitude', 'Latitude'])
        circlegenerateddf['Radius'] = radius

    else:
        rejeectedcoordinates = []
        uncoverpoints = finduncoveredpoints(polygon,circle_positions, radius_deg)

        numberofcircle = len(circle_positions)

        if data['isChecked'] == True: #after generating a random circle to start the process of finding uncovered points in the polygon
            max_time_limit = 30
            start_time = time.time()
            while finduncoveredpoints(polygon,circle_positions, radius_deg) != []: #checks if there are still uncovered points
                if time.time() - start_time > max_time_limit:
                    break
                coordinates = random.choice(uncoverpoints) #picks random uncovered coordinates
                new_circle = {'x':coordinates[0], 'y':coordinates[1], 'radius':radius_deg}
                center = [coordinates[0], coordinates[1]]

                #getting the coordinates of the circumference and make it as a "polygon variable"
                circle_geojson = circle_to_geojson(coordinates[0], coordinates[1], radius_deg)
                #making it a complete polygon by mentioning the first coordinates at the end of the list
                coordinates = circle_geojson['coordinates'][0]
                coordinatesfirst = coordinates[0]
                coordinates.append(coordinatesfirst)

                if circle_inside_polygon(coordinates) and coordinates not in rejeectedcoordinates:
                    if circles_overlap(new_circle, existing_circles):
                        if finduncoveredpoints(polygon,circle_positions, radius_deg) == []:
                            break
                        circle_positions.append({"center":center,"coordinates":coordinates}) #store coordinates
                        existing_circles.append(new_circle)
                        numberofcircle += 1
                    else:
                        rejeectedcoordinates.append(coordinates)
                if time.time() - start_time > max_time_limit:
                    break
        if numberofcircle == len(circle_positions):
            notify = f'Recommended number of circle: {numberofcircle}'
        content += f'Number of Circles: {numberofcircle}\n'

        for positions in circle_positions: #for many circles
            number += 1
            longitude = positions['center'][0]
            longitude = '{:.6f}'.format(longitude)

            latitude = positions['center'][1]
            latitude = '{:.6f}'.format(latitude)

            content += f'\nCircle {number}\nCenter:\nLongitude: {longitude} Latitude: {latitude}\n'
            circlecoordinateslist.append(positions['center'])

        if maxcircle == False and notify == '': #if the circles can fit in the polygon
            notify = ''
        else: #if it reaches the maximum number of circles to fit in the polygon
            notify = f'Max number of circles: {number}'
    
    htmlcontent = content.replace('\n', '<br>')

    circlegenerateddf = pd.DataFrame(circlecoordinateslist, columns=['Longitude', 'Latitude'])
    circlegenerateddf['Radius'] = radius


    return json.dumps({'area': printarea, 'content':htmlcontent, 'projected_coords':coordinateslist, 'circle_coords': circlecoordinateslist
                       , 'notify': notify})


@app.route('/send_tempfile', methods=['POST'])
def send_tempfile(): #to send the output of the text container and the csv or excel sheet which will be in a zip
    def create_temp_dirs():
        global UPLOAD_FOLDER
        global OUTPUT_FOLDER
        global ZIP_FOLDER
        OUTPUT_FOLDER = tempfile.TemporaryDirectory()
        ZIP_FOLDER = tempfile.TemporaryDirectory()
        UPLOAD_FOLDER = tempfile.TemporaryDirectory()


    def deletefiles():
        UPLOAD_FOLDER.cleanup()
        OUTPUT_FOLDER.cleanup()
        ZIP_FOLDER.cleanup()
        create_temp_dirs()

    deletefiles()

    txt_file = open(os.path.join(OUTPUT_FOLDER.name, 'output.txt'), 'w', encoding='utf-8') #It is the output in the output container
    if searchcontent != '':
        txt_file.write(searchcontent)
    elif replacecontent != '':
        txt_file.write(replacecontent)
    else:
        print(content)
        txt_file.write(content)


    txt_file.close()

    if employeedict != {} and projectplace != []: #to store employee details in the csv file
        print(employeedict)
        csv_file = open(os.path.join(OUTPUT_FOLDER.name, f"{filename.rstrip('.xlsx')}_out.csv"), 'w', newline='')
        fields = ['BadgeID', 'EmpName', 'ClockDate', 'ClockTime', 'Latitude', 'Longitude', 'ProjectCode', 'Status']
        writer = csv.DictWriter(csv_file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(employeedict)
        csv_file.close()

        zipfilepath = os.path.join(ZIP_FOLDER.name, f'{filename.rstrip(".xlsx")}.zip')
        print(zipfilepath)

        with ZipFile(zipfilepath, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(OUTPUT_FOLDER.name):
                for file in files:
                    filepath = os.path.join(root, file)
                    zip_file.write(filepath, os.path.relpath(filepath, OUTPUT_FOLDER.name))

        response = send_file(zipfilepath, as_attachment=True, download_name=f'{filename.rstrip(".xlsx")}.zip')


    elif (coordinatesdf.empty is not True and employeedict != {}) or circlegenerateddf.empty is not True: #to store the coordinates of the generated circle into a excel spreadsheet
        if coordinatesdf.empty is not True and employeedict != {}:    
            xlsx_file_path = os.path.join(OUTPUT_FOLDER.name, f"{otherfilename.rstrip('.xlsx')}_out.xlsx")
        else:
            xlsx_file_path = os.path.join(OUTPUT_FOLDER.name, f"generated_coords.xlsx")

        with open(xlsx_file_path, 'wb') as xlsx_file:
            if circlegenerateddf.empty is not True:
                circlegenerateddf.to_excel(xlsx_file, index=False)
            else:    
                coordinatesdf.to_excel(xlsx_file, index=False)


        if coordinatesdf.empty is not True and employeedict != {}:    
            zipfilepath = os.path.join(ZIP_FOLDER.name, f'{otherfilename.rstrip(".xlsx")}.zip')
        else:
            zipfilepath = os.path.join(ZIP_FOLDER.name, f'generated.zip')

        with ZipFile(zipfilepath, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(OUTPUT_FOLDER.name):
                for file in files:
                    filepath = os.path.join(root, file)
                    zip_file.write(filepath, os.path.relpath(filepath, OUTPUT_FOLDER.name))

        if coordinatesdf.empty is not True and employeedict != {}:    
            response = send_file(zipfilepath, as_attachment=True, download_name=f'{otherfilename.rstrip(".xlsx")}.zip')
        else:
            response = send_file(zipfilepath, as_attachment=True, download_name=f'generated.zip')

    else:
        response = send_file(os.path.join(OUTPUT_FOLDER.name, 'output.txt'), as_attachment=True, download_name=f'output.txt')

    return response

@app.route('/upload', methods=['POST']) #when the user upload the files, they will be added into the UPLOAD_FOLDER
def upload_files():
    deletefiles()
    if 'files' not in request.files:
        return jsonify({'error': 'No files part'}), 400

    files = request.files.getlist('files')
    for file in files:
        if file:
            filename = file.filename
            file.save(os.path.join(UPLOAD_FOLDER.name, filename))

    chooseemployee = []

    for inputfiles in os.listdir(UPLOAD_FOLDER.name): #this is to check if that file is a clock record file (incase user wants to access the past files)
        filepath = f'{UPLOAD_FOLDER.name}/{inputfiles}'
        wb = openpyxl.load_workbook(filepath)
        sheetnames = wb.sheetnames
        for sheet in sheetnames:
            inputfiledf = pd.read_excel(filepath, sheet_name=sheet)
            if 'EmployeeCodeName' in inputfiledf.columns: 
                for latitude, longitude, employeenamecode in zip(inputfiledf['Latitude'], inputfiledf['Longitude'], inputfiledf['EmployeeCodeName']):
                    if (longitude != None and latitude != None) and (longitude != 0 and latitude != 0):
                        chooseemployee.append(employeenamecode)

    chooseemployee = list(set(chooseemployee)) #it will be stored in the drop down that allows the user to choose the employee before processing

    return jsonify({
        'message': 'Files uploaded successfully',
        'chooseemployee': chooseemployee
    }), 200

@app.route('/searchoutput', methods=['POST'])
def searchoutput(): #this is to filter the output which is including the map and the details in the output container. They get to filter the name and the date. 
    global searchcontent
    global radius
    data = request.get_json()
    findemployee = data['employee'] #get the selected employee name from the dropdown box
    finddate = data['finddate'] #get the selected date from the datepicker
    coordlist = []
    projectplacecoords = []

    searchcontent = ''

    if (findemployee):
        if findemployee != 'allemployee':
            searchcontent += f'Employee: {findemployee}\n'
    if (finddate):
        finddate = datetime.strptime(finddate, '%Y-%m-%d')
        finddate = finddate.date()
        searchcontent += f'Date: {finddate}\n'


    if filteremployee[0]['shapetype'] == 'circle':
        radius = data["radius"]
        searchcontent += f'Radius: {radius}\n'

    canbefound = False
    number = 0
    for i in range(len(filteremployee)):
        shapetype = filteremployee[i]['shapetype']

        if filteremployee[i]['EmpName'] == findemployee and filteremployee[i]['ClockDate'] == finddate: #if they are the same then ot will produce this
            number += 1
            if filteremployee[i]['shapetype'] == 'circle':
                searchcontent += f"\nCircle {number}\nEmployee Name: {filteremployee[i]['EmpName']}\nBadge ID: {filteremployee[i]['BadgeID']}\nDate: {filteremployee[i]['ClockDate'].strftime('%d/%m/%Y')} {filteremployee[i]['ClockTime']}\nProject Site: {filteremployee[i]['closestlocation']}\nLocation Name: {filteremployee[i]['locationname']}\nCenter:\nLongitude: {filteremployee[i]['Center'][0]} Latitude: {filteremployee[i]['Center'][1]}\nFile Name: {filteremployee[i]['inputfile']}\nSheet Name: {filteremployee[i]['sheetname']}\n"
            else:
                searchcontent += f"\nPin {number}\nEmployee Name: {filteremployee[i]['EmpName']}\nBadge ID: {filteremployee[i]['BadgeID']}\nDate: {filteremployee[i]['ClockDate'].strftime('%d/%m/%Y')} {filteremployee[i]['ClockTime']}\nProject Site: {filteremployee[i]['closestlocation']}\nLocation Name: {filteremployee[i]['locationname']}\nCenter:\nLongitude: {filteremployee[i]['Center'][0]} Latitude: {filteremployee[i]['Center'][1]}\nFile Name: {filteremployee[i]['inputfile']}\nSheet Name: {filteremployee[i]['sheetname']}\n"
            coordlist.append(filteremployee[i]['Center'])
            shapetype = filteremployee[i]['shapetype']

            canbefound = True

        #to prevent producing the one that cannot be found
        elif (filteremployee[i]['EmpName'] == findemployee and finddate == '') or ((findemployee == '' or findemployee == 'allemployee') and filteremployee[i]['ClockDate'] == finddate):
            number += 1
            if filteremployee[i]['shapetype'] == 'circle':
                searchcontent += f"\nCircle {number}\nEmployee Name: {filteremployee[i]['EmpName']}\nBadge ID: {filteremployee[i]['BadgeID']}\nDate: {filteremployee[i]['ClockDate'].strftime('%d/%m/%Y')} {filteremployee[i]['ClockTime']}\nProject Site: {filteremployee[i]['closestlocation']}\nLocation Name: {filteremployee[i]['locationname']}\nCenter:\nLongitude: {filteremployee[i]['Center'][0]} Latitude: {filteremployee[i]['Center'][1]}\nFile Name: {filteremployee[i]['inputfile']}\nSheet Name: {filteremployee[i]['sheetname']}\n"
            else:
                searchcontent += f"\nPin {number}\nEmployee Name: {filteremployee[i]['EmpName']}\nBadge ID: {filteremployee[i]['BadgeID']}\nDate: {filteremployee[i]['ClockDate'].strftime('%d/%m/%Y')} {filteremployee[i]['ClockTime']}\nProject Site: {filteremployee[i]['closestlocation']}\nLocation Name: {filteremployee[i]['locationname']}\nCenter:\nLongitude: {filteremployee[i]['Center'][0]} Latitude: {filteremployee[i]['Center'][1]}\nFile Name: {filteremployee[i]['inputfile']}\nSheet Name: {filteremployee[i]['sheetname']}\n"
            coordlist.append(filteremployee[i]['Center'])
            shapetype = filteremployee[i]['shapetype']

            canbefound = True

    if canbefound == False:
        searchcontent = '<b>Not Found</b>'

    if projectplace != []:
        searchcontent += '\n\nPROJECT SITES\n\n'
        searchcontent += '=' * 50
        for p in range(len(projectplace)):
            number += 1
            searchcontent += f"\nCircle {number}\n{projectplace[p]['Location']}\nAddress: {projectplace[p]['Address']}\nLongitude: {projectplace[p]['Center'][0]} Latitude: {projectplace[p]['Center'][1]}\n"
            projectplacecoords.append(projectplace[p]['Center'])
            shapetype = projectplace[p]['shapetype']


    if findemployee == 'allemployee' and finddate == '':
        if projectplace == []:
            for i in range(len(filteremployee)):
                shapetype = filteremployee[i]['shapetype']
                coordlist.append(filteremployee[i]['Center'])
            searchcontent = content
        else:
            for i in range(len(filteremployee)):
                coordlist.append(filteremployee[i]['Center'])
            shapetype = 'circle'
            searchcontent = content


    htmlcontent = searchcontent.replace('\n','<br>') #to put it in the innerhtml in the text container
    return json.dumps({'findcoords': coordlist, 'content':htmlcontent, 'shapetype': shapetype,'projectplacecoords': projectplacecoords})


@app.route('/process', methods=['POST']) #this is the process for the upload coords
def coordprocess():
    global filteremployee
    global coordinatesdf
    global employeedict
    global filename
    global otherfilename
    global replacecontent
    global searchcontent
    global content

    content = projectplacecontent

    

    replacecontent = ''
    searchcontent = ''
    otherfilename = ''
    employeedict = {}
    coordinatesdf = pd.DataFrame()
    filename = ''

    data = request.get_json()
    listofdir = os.listdir(UPLOAD_FOLDER.name)
    if len(listofdir) == 0 and os.path.exists(clockfile) == False:
        return jsonify({'error': 'No data or files provided'}), 400
    circlecoordinateslist = []
    number = 0
    def circles_overlap(circle1, circle2): #check if it overlapped more than its radius
        if circle2 != []:
            checkall = False
            for circle in circle2:
                distance = Point(circle1['x'], circle1['y']).distance(Point(circle['x'], circle['y'])) #check the distance of the two circle center point
                if distance < (circle1['radius'])*0.5: #if the distance is smaller thab the radius it will not be accepted because it overlap alot
                    return False
                else:
                    checkall = True
            
            if checkall == True:
                return True
            
        else: #so if the list is empty it will put the first circle then compare the rest if needed
            return True
        
    def circles_overlap_finder(circle1, circle2): #check if it overlapped more than its radius
        if circle2 != []:
            for circle in circle2:
                distance = Point(circle1['x'], circle1['y']).distance(Point(circle['x'], circle['y'])) #check the distance of the two circle center point
                if distance < (circle1['radius'])*0.5: #if the distance is smaller thab the radius it will not be accepted because it overlap alot
                    return circle

    filteremployee = []
    filteremployeename = []
    getdate = []
    radiuslist = []

    checkemployee = False
    checklocation = False
    checkmobileclock = False
    checkradius = False


    if os.path.exists(clockfile) == False: #checks if the directory is there or not
        for file in os.listdir(UPLOAD_FOLDER.name):
            filepath = f'{UPLOAD_FOLDER.name}/{file}'
            wb = openpyxl.load_workbook(filepath)
            sheetnames = wb.sheetnames
            for sheet in sheetnames:
                inputfiledf = pd.read_excel(filepath, sheet_name=sheet)
                if 'EmployeeCodeName' in inputfiledf.columns:
                    checkemployee = True
                if 'Radius' in inputfiledf.columns:
                    checkradius = True
    else:
        checkemployee = True


    if projectplace != []:
        checklocation = True

    #just checks if it has occur to make it easier
    if data['coordradius'] != '' and checklocation == False:
        radius = float(data['coordradius'])
        content += 'Circles\n'
        content += f'Circle Radius: {radius}m\n'
    elif checklocation and data['coordradius'] == '':
        radius = 100
        content += 'Project Site Area\n'
        content += f'Circle Radius: 100m\n'
    elif checklocation and data['coordradius'] != '':
        radius = float(data['coordradius'])
        content += 'Project Site Area\n'
        content += f'Circle Radius: {radius}m\n'
    elif checkradius and data['coordradius'] == '':
        content += 'Circles\n'
    else:
        content += 'Pins\n'

    if checkemployee and checklocation:
        checkmobileclock = True

    #to make sure that it is not empty and if it is it will be as allemployee
    chosen_employee_value = data.get('chosenemployeevalue', None)
    if not chosen_employee_value:
        data['chosenemployeevalue'] = 'allemployee'
    
    #for clockfile file to be processed
    if os.path.exists(clockfile) and len(os.listdir(UPLOAD_FOLDER.name)) == 0:
        wb = openpyxl.load_workbook(clockfile)
        clocksplit = clockfile.split('\\')

        relativepath = clocksplit[-1]
        filename = relativepath
        sheetnames = wb.sheetnames
        for sheet in sheetnames:
            inputfiledf = pd.read_excel(clockfile, sheet_name=sheet)
            for latitude, longitude, employeenamecode, clockdate, clocktime, badgeno, locationname in zip(inputfiledf['Latitude'], inputfiledf['Longitude'], inputfiledf['EmployeeCodeName'], inputfiledf['ClockDate'], inputfiledf['ClockTime'], inputfiledf['BadgeNo'], inputfiledf['LocationName']):
                if (longitude != None and latitude != None) and (longitude != 0 and latitude != 0):
                    source_crs = pyproj.Proj(init='epsg:3857')
                    geodetic = Proj(init='epsg:4326')  # WGS84
                    center_x, center_y = pyproj.transform(source_crs, geodetic, latitude, longitude)
                    radius_deg = pyproj.transform(source_crs, geodetic, radius+1.34, 0)[0] - center_x
                    
                    closestdistance = pyproj.transform(source_crs, geodetic, 1000000, 0)[0] - center_x
                    closestlocation = ''
                    found = False

                    #this is to get the closest project site they are close to
                    for p, indexproj in zip(projectplacecoords, range(len(projectplace))) :
                        distance = Point(p[0], p[1]).distance(Point(longitude, latitude))
                        if distance <= closestdistance:
                            closestdistance = distance
                            closestlocation = projectplace[indexproj]['Location']

                    if employeenamecode == data['chosenemployeevalue'] or data['chosenemployeevalue'] == 'allemployee':
                        number += 1
                        center = [longitude, latitude]
                        circlecoordinateslist.append(center)
                        filteremployee.append({'Number': number, 'EmpName': employeenamecode, 'BadgeID': badgeno,'Longitude':longitude,'Latitude':latitude,
                                                        'Center':center,'ClockDate': clockdate.date(),'ClockTime' : clocktime,'sheetname': sheet, 'inputfile': relativepath, 'shapetype': 'pin','closestlocation' : closestlocation, 'locationname': locationname})
                        filteremployeename.append(employeenamecode)

                        
                        content += f'\nPin {number}\nEmployee Name: {employeenamecode}\nBadge ID: {badgeno}\nDate: {clockdate.date().strftime("%d/%m/%Y")} {clocktime}\nProject Site: {closestlocation}\nLocation Name: {locationname}\nCenter:\nLongitude: {longitude} Latitude: {latitude}\nFile Name: {relativepath}\nSheet Name: {sheet}\n'
        
    #for the files that they input
    if len(os.listdir(UPLOAD_FOLDER.name)) > 0:
        for inputfiles in os.listdir(UPLOAD_FOLDER.name):

            existing_circles = []
            employeedictlist = []
            yesemployee = False
            notspecial = False
            withradius = False
    
            content += '\n' + '=' * 50 + '\n\n'
            content += f'File Name: {inputfiles}\n'
            content += '\n' + '=' * 50 
            filepath = f'{UPLOAD_FOLDER.name}/{inputfiles}'
            wb = openpyxl.load_workbook(filepath)
            sheetnames = wb.sheetnames
            for sheet in sheetnames:
                inputfiledf = pd.read_excel(filepath, sheet_name=sheet)

                #check if its the clock records
                if 'EmployeeCodeName' in inputfiledf.columns:
                    filename = inputfiles

                    for latitude, longitude, employeenamecode, clockdate, clocktime, badgeno, locationname in zip(inputfiledf['Latitude'], inputfiledf['Longitude'], inputfiledf['EmployeeCodeName'], inputfiledf['ClockDate'], inputfiledf['ClockTime'], inputfiledf['BadgeNo'], inputfiledf['LocationName']):

                        if (longitude != None and latitude != None) and (longitude != 0 and latitude != 0):

                            
                            source_crs = pyproj.Proj(init='epsg:3857')
                            geodetic = Proj(init='epsg:4326')  # WGS84
                            center_x, center_y = pyproj.transform(source_crs, geodetic, latitude, longitude)
                            radius_deg = pyproj.transform(source_crs, geodetic, radius+1.34, 0)[0] - center_x
                            
                            closestdistance = pyproj.transform(source_crs, geodetic, 1000000, 0)[0] - center_x
                            closestlocation = ''
                            found = False

                            for p, indexproj in zip(projectplacecoords, range(len(projectplace))) :
                                distance = Point(p[0], p[1]).distance(Point(longitude, latitude))
                                if distance <= closestdistance:
                                    closestdistance = distance
                                    closestlocation = projectplace[indexproj]['Location']

                            print(data['chosenemployeevalue'])

                            if employeenamecode == data['chosenemployeevalue'] or data['chosenemployeevalue'] == 'allemployee':
                                number += 1
                                center = [longitude, latitude]
                                circlecoordinateslist.append(center)
                                filteremployee.append({'Number': number, 'EmpName': employeenamecode, 'BadgeID': badgeno,'Longitude':longitude,'Latitude':latitude,
                                                                'Center':center,'ClockDate': clockdate.date(),'ClockTime' : clocktime,'sheetname': sheet, 'inputfile': inputfiles, 'shapetype': 'pin','closestlocation' : closestlocation, 'locationname': locationname})
                                filteremployeename.append(employeenamecode)

                                
                                content += f'\nPin {number}\nEmployee Name: {employeenamecode}\nBadge ID: {badgeno}\nDate: {clockdate.date().strftime("%d/%m/%Y")} {clocktime}\nProject Site: {closestlocation}\nLocation Name: {locationname}\nCenter:\nLongitude: {longitude} Latitude: {latitude}\nFile Name: {inputfiles}\nSheet Name: {sheet}\n'

                #check if the file contains radius
                elif 'Radius' in inputfiledf.columns:
                    content = ''
                    print('here')
                    otherfilename = inputfiles
                    content += '\n' + '-' * 50 + '\n\n'
                    content += f'Sheet Name: {sheet}\n'
                    for latitude, longitude, fileradius in zip(inputfiledf['Latitude'], inputfiledf['Longitude'], inputfiledf['Radius']):
                        if (longitude != None and latitude != None) and (longitude != 0 and latitude != 0):
                            if data['coordradius'] == '':
                                circlecoordinateslist.append([longitude, latitude])
                                radiuslist.append(fileradius)
                                number += 1
                                content += f'\nCircle {number}\nCenter:\nLongitude: {longitude} Latitude: {latitude}\nRadius: {fileradius}\n'
                                print(radiuslist)
                            else:
                                center = [longitude, latitude]
                                circlecoordinateslist.append(center)
                                number += 1
                                content += f'\nCircle {number}\nCenter:\nLongitude: {longitude} Latitude: {latitude}\n'
                    
                    withradius =True

                #check if the file has longitude or latitude
                else:
                    content = ''
                    otherfilename = inputfiles
                    content += '\n' + '-' * 50 + '\n\n'
                    content += f'Sheet Name: {sheet}\n'
                    for latitude, longitude in zip(inputfiledf['Latitude'], inputfiledf['Longitude']):
                        if (longitude != None and latitude != None) and (longitude != 0 and latitude != 0):

                            if data['coordradius'] != '':
                                if data['overlapchecked'] == True:

                                    source_crs = pyproj.Proj(init='epsg:3857')
                                    geodetic = Proj(init='epsg:4326')  # WGS84
                                    center_x, center_y = pyproj.transform(source_crs, geodetic, latitude, longitude)
    
                                    radius_deg = pyproj.transform(source_crs, geodetic, radius+1.34, 0)[0] - center_x

                                    new_circle = {'x': longitude, 'y': latitude, 'radius': radius_deg}

                                    if circles_overlap(new_circle, existing_circles):

                                        center = [longitude, latitude]
                                        existing_circles.append(new_circle)
                                        circlecoordinateslist.append(center)
                                        number += 1
                                        content += f'\nCircle {number}\nCenter:\nLongitude: {longitude} Latitude: {latitude}\n'
                                else:
                                    center = [longitude, latitude]
                                    circlecoordinateslist.append(center)
                                    number += 1
                                    content += f'\nCircle {number}\nCenter:\nLongitude: {longitude} Latitude: {latitude}\n'
                            else:
                                center = [longitude, latitude]
                                circlecoordinateslist.append(center)
                                number += 1
                                content += f'\nPin {number}\nLongitude: {longitude} Latitude: {latitude}\n'
                    notspecial = True
            if yesemployee == True:
                for sheet in sheetnames:
                    content += '\n' + '-' * 50 + '\n\n'
                    content += f'Sheet Name: {sheet}\n'
                    if employeedictlist != []:
                        for i in range(len(employeedictlist)):
                            if employeedictlist[i]['sheetname'] == sheet:
                                content += f'\nCircle {employeedictlist[i]["Number"]}\nEmployee Name:\n'
                                for j in range(len(employeedictlist[i]['EmpName'])):

                                    content += f'{employeedictlist[i]["EmpName"][j]}\n'
                                content += f'Center:\nLongitude: {employeedictlist[i]["Center"][0]} Latitude: {employeedictlist[i]["Center"][1]}\n'


    if filteremployee != [] and projectplacecoords != []:

        for i in range(len(filteremployee)):
            longitude = filteremployee[i]['Longitude']
            latitude = filteremployee[i]['Latitude']
            source_crs = pyproj.Proj(init='epsg:3857')
            geodetic = Proj(init='epsg:4326')  # WGS84
            center_x, center_y = pyproj.transform(source_crs, geodetic, latitude, longitude)
            radius_deg = pyproj.transform(source_crs, geodetic, radius+1.34, 0)[0] - center_x
            outerradius_deg = pyproj.transform(source_crs, geodetic, radius+21.34, 0)[0] - center_x

            closestdistance = pyproj.transform(source_crs, geodetic, 1000000, 0)[0] - center_x
            closestlocation = ''
            found = False

            #to check if the user is in range and produce the status Y, Y1. N 
            #Y - within 100m
            #Y1 - within 120m
            #N - outside the range

            for p, indexproj in zip(projectplacecoords, range(len(projectplace))) :
                distance = Point(p[0], p[1]).distance(Point(longitude, latitude))
                if distance <= radius_deg:
                    filteremployee[i]['ProjectCode'] = projectplace[indexproj]['Location']
                    filteremployee[i]['Status'] = 'Y'
                    found = True
                    break
                elif distance <= outerradius_deg:
                    filteremployee[i]['ProjectCode'] = projectplace[indexproj]['Location']
                    filteremployee[i]['Status'] = 'Y1'
                    found = True
                    break
                else:
                    if distance <= closestdistance:
                        closestdistance = distance
                        closestlocation = projectplace[indexproj]['Location']
                    filteremployee[i]['Status'] = 'N'
            
            if found == False:
                filteremployee[i]['ProjectCode'] = closestlocation


        #making the output

        employeedf = pd.DataFrame.from_dict(filteremployee)

        employeedf = employeedf[['BadgeID', 'EmpName', 'ClockDate', 'ClockTime', 'Latitude', 'Longitude', 'ProjectCode', 'Status']]
        
        employeedict = employeedf.to_dict('records')


        #upload the filters into the webpage
        for i in range(len(filteremployee)):
            getdate.append(filteremployee[i]['ClockDate'])

        getdate = list(set(getdate))
        getdate.sort()
        
        startdate = str(getdate[0])
        enddate = str(getdate[-1])

        htmlcontent = content.replace('\n','<br>')

        if data['coordradius'] != '':
            notify = f'Number of Circles: {number}'
        else: 
            notify = f'Number of Pins: {number}'

        filteremployeename = list(set(filteremployeename))

        if projectplacecoords != []:

            return json.dumps({'circle_coords': circlecoordinateslist,'projectplacecoords': projectplacecoords ,'content':htmlcontent, 'notify':notify, 'employeenames':filteremployeename, 'startdate': startdate, 'enddate': enddate, 'checkmobileclock':checkmobileclock, 'locationlist': locationlist})
        else:

            return json.dumps({'circle_coords': circlecoordinateslist, 'content':htmlcontent, 'notify':notify, 'employeenames':filteremployeename, 'startdate': startdate, 'enddate': enddate, 'checkmobileclock':checkmobileclock, 'locationlist': locationlist})
    
    elif filteremployee !=  []:
        for i in range(len(filteremployee)):
            getdate.append(filteremployee[i]['ClockDate'])
        
        getdate = list(set(getdate))
        getdate.sort()
        
        startdate = str(getdate[0])
        enddate = str(getdate[-1])
        
        htmlcontent = content.replace('\n','<br>')

        if data['coordradius'] != '':
            notify = f'Number of Circles: {number}'
        else: 
            notify = f'Number of Pins: {number}'


        filteremployeename = list(set(filteremployeename))


        if projectplacecoords == []:
            print('here3')
            return json.dumps({'circle_coords': circlecoordinateslist, 'content':htmlcontent, 'notify':notify, 'employeenames':filteremployeename, 'startdate': startdate, 'enddate': enddate, 'checkmobileclock':checkmobileclock, 'locationlist': locationlist})

    
    else:

    
        coordinatesdf = pd.DataFrame(circlecoordinateslist, columns=['Longitude', 'Latitude'])

        if radiuslist != []:
            coordinatesdf['Radius'] = radiuslist
        elif data['coordradius'] != '':
            coordinatesdf['Radius'] = radius

        htmlcontent = content.replace('\n','<br>')

        if data['coordradius'] != '' or radiuslist != []:
            notify = f'Number of Circles: {number}'
        else: 
            notify = f'Number of Pins: {number}'


        if radiuslist != []:
            print('here4')
            return json.dumps({'circle_coords': circlecoordinateslist, 'radiuslist': radiuslist, 'content':htmlcontent, 'notify': notify, 'withradius': withradius})
        else:
            print('here5')
            return json.dumps({'circle_coords': circlecoordinateslist, 'content':htmlcontent, 'notify':notify, 'checkmobileclock':checkmobileclock, 'locationlist': locationlist, 'notspecial':notspecial, 'withradius':withradius})
        




@app.route('/editoutput',  methods=['POST']) #to edit the gerated circles 
def edit_output():
    global replacecontent
    global circlegenerateddf
    replacecontent = ''
    data = request.get_json()

    replacecontent = data['htmlcontent']
    print(data['add'])
    if data['add']: #when adding, it will add on to the output text and the dataframe that will be sent as a csv file or an excel
        replacecontent = replacecontent.replace('<br>', '\n')
        somestring = re.findall('Circle \d+', replacecontent)
        if somestring != []:
            somestring = [int(s.replace('Circle ', '')) for s in somestring]
            maxnumber = max(somestring)
        else:
            maxnumber = 0

        circlecoordinates = json.loads(data['circlecoordinates'])
        replacecontent += f'\nCircle {maxnumber + 1}\nCenter:\nLongitude: {circlecoordinates[0]} Latitude: {circlecoordinates[1]}\n'
        htmlcontent = replacecontent.replace('\n', '<br>')

        newcirclelist = [circlecoordinates[0], circlecoordinates[1], round(float(data['thisradius']),2)]
        
        if circlegenerateddf.empty is not True:
            circlegenerateddf.loc[circlegenerateddf.index.max()+ 1] = newcirclelist
        if circlegenerateddf.empty:
            circlegenerateddf.loc[0] = newcirclelist


        return json.dumps({'content': htmlcontent, 'number': maxnumber})
    else: #when deleting, it will delete the output of the marker that u deleted and remove the row in the dataframe
        somestring = replacecontent.replace('<br>', '\n')

        somestring = somestring.replace('<span class="highlight">', '')
        somestring = somestring.replace('</span>', '')

        deleteditem = json.loads(data['deleteditem'])


        for item in deleteditem:
            item = int(item.replace('Pin ', ''))
            pattern = re.compile(rf"\nCircle {item}\nCenter:\nLongitude: [0-9.]+ Latitude: [0-9.]+\n")

            circlegenerateddf.drop(index=item - 1, inplace=True)
            

            somestring = re.sub(pattern, "", somestring)
        replacecontent = somestring
        htmlcontent = somestring.replace('\n', '<br>')
   

        return json.dumps({'content': htmlcontent})


def open_webbrowser():
    webbrowser.open_new('http://127.0.0.1:8000')

if __name__ == '__main__':
    
    open_webbrowser()
    app.run(use_reloader = False, port=8000)

